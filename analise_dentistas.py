

import requests
import pandas as pd
from datetime import datetime
import sys
import time
import base64

# =========================================================
# CONFIGURAÇÃO - PREENCHA AQUI
# =========================================================

BASE_URL = "https://api.clinicorp.com/rest/v1"
ENDPOINT = "/estimates/list"

# Dados confirmados no Swagger (estimates/list)
SUBSCRIBER_ID = ""   # parâmetro "subscriber_id"
CLINIC_ID = ""           # parâmetro "clinic_id" (opcional - deixe "" se não usar multiclínicas)

# Autenticação Basic Auth (usuário e senha da Clinicorp)
USERNAME = ""
PASSWORD = ""

# Nome do arquivo de saída
OUTPUT_FILE = "analise_faturamento_dentistas.xlsx"


def gerar_basic_auth(usuario, senha):
    """Gera o header Basic Auth a partir de usuário e senha."""
    credenciais = f"{usuario}:{senha}"
    return base64.b64encode(credenciais.encode("utf-8")).decode("utf-8")


# =========================================================
# FUNÇÕES
# =========================================================

def pedir_periodo():
    """Pergunta ao usuário o período de análise (data início e fim)."""
    print("=== Análise de Faturamento por Dentista ===\n")
    while True:
        data_inicio = input("Data inicial (formato AAAA-MM-DD): ").strip()
        data_fim = input("Data final (formato AAAA-MM-DD): ").strip()
        try:
            di = datetime.strptime(data_inicio, "%Y-%m-%d")
            df = datetime.strptime(data_fim, "%Y-%m-%d")
            if di > df:
                print("Data inicial não pode ser depois da data final. Tente de novo.\n")
                continue
            return data_inicio, data_fim
        except ValueError:
            print("Formato inválido. Use AAAA-MM-DD (ex: 2026-01-01).\n")


def buscar_estimates(data_inicio, data_fim):
    """
    Faz a chamada à API estimates/list para o período informado.
    A documentação não indicou paginação explícita; o código já trata
    o caso de a API devolver tudo de uma vez (mais comum) e também o
    caso de vir limitada, com um aviso para você conferir manualmente.
    """
    auth_b64 = gerar_basic_auth(USERNAME, PASSWORD)

    params = {
        "subscriber_id": SUBSCRIBER_ID,
        "from": data_inicio,
        "to": data_fim,
    }
    if CLINIC_ID:
        params["clinic_id"] = CLINIC_ID

    headers = {
        "accept": "application/json",
        "Authorization": f"Basic {auth_b64}",
    }

    print("Buscando dados na API...")

    try:
        resp = requests.get(
            f"{BASE_URL}{ENDPOINT}",
            params=params,
            headers=headers,
            timeout=30,
        )
    except requests.exceptions.RequestException as e:
        print(f"Erro de conexão: {e}")
        sys.exit(1)

    if resp.status_code == 401:
        print("Erro 401: credenciais inválidas. Confira USERNAME/PASSWORD.")
        sys.exit(1)
    elif resp.status_code == 429:
        print("Limite de requisições atingido (429). Tente novamente mais tarde.")
        sys.exit(1)
    elif resp.status_code != 200:
        print(f"Erro {resp.status_code}: {resp.text[:300]}")
        sys.exit(1)

    dados = resp.json()

    print(f"Total de registros (orçamentos) coletados: {len(dados)}\n")

    if len(dados) >= 500:
        print(
            "Aviso: você recebeu 500+ registros em uma única chamada — esse é o\n"
            "limite por hora confirmado nos headers da API (x-ratelimit-limit-hour).\n"
            "Se esperava mais dados que isso, talvez precise dividir o período em\n"
            "janelas menores (ex: mês a mês) e rodar o script mais de uma vez.\n"
        )

    return dados


def processar_dados(registros):
    """
    Transforma a lista de orçamentos em tabelas de análise.

    Estrutura real confirmada via API:
    - Cada orçamento tem um ProfessionalName/ProfessionalId "dono" do orçamento
    - Cada item dentro de ProcedureList tem seu próprio DentistName/Dentist_PersonId,
      que pode ser diferente do profissional do orçamento (ex: procedimento
      encaminhado para outro dentista dentro do mesmo orçamento)
    - Itens com "Deleted": "X" foram excluídos e não devem contar no faturamento
    - O valor real cobrado por procedimento é FinalAmount (não Amount,
      que é o valor de tabela antes de ajustes)
    """
    linhas = []

    for est in registros:
        orc_prof_id = est.get("ProfessionalId")
        orc_prof_nome = est.get("ProfessionalName")
        status_orcamento = est.get("Status")
        data = est.get("Date")
        paciente = est.get("PatientName")
        treatment_id = est.get("TreatmentId")

        procedimentos = est.get("ProcedureList", [])

        if not procedimentos:
            linhas.append({
                "TreatmentId": treatment_id,
                "Data": data,
                "Paciente": paciente,
                "Status_Orcamento": status_orcamento,
                "Profissional_Orcamento": orc_prof_nome,
                "ProfessionalId_Orcamento": orc_prof_id,
                "Procedimento": None,
                "Dente": None,
                "Profissional_Executante": orc_prof_nome,
                "Dentist_PersonId_Executante": orc_prof_id,
                "Valor_Tabela": est.get("Amount", 0),
                "Valor_Final": est.get("Amount", 0),
                "Deletado": False,
            })
            continue

        for proc in procedimentos:
            deletado = proc.get("Deleted") == "X"
            linhas.append({
                "TreatmentId": treatment_id,
                "Data": data,
                "Paciente": paciente,
                "Status_Orcamento": status_orcamento,
                "Profissional_Orcamento": orc_prof_nome,
                "ProfessionalId_Orcamento": orc_prof_id,
                "Procedimento": proc.get("OperationDescription"),
                "Dente": proc.get("Tooth"),
                "Profissional_Executante": proc.get("DentistName"),
                "Dentist_PersonId_Executante": proc.get("Dentist_PersonId"),
                "Valor_Tabela": proc.get("Amount", 0),
                "Valor_Final": proc.get("FinalAmount", proc.get("Amount", 0)),
                "Deletado": deletado,
            })

    df_detalhado = pd.DataFrame(linhas)

    if df_detalhado.empty:
        print("Nenhum dado retornado para o período. Verifique os filtros.")
        sys.exit(0)

    # Remove procedimentos marcados como deletados do cálculo de faturamento
    df_ativos = df_detalhado[~df_detalhado["Deletado"]].copy()

    status_unicos = sorted(df_detalhado["Status_Orcamento"].dropna().unique().tolist())
    print(f"Status de orçamento encontrados no período: {status_unicos}")
    print(
        f"{df_detalhado['Deletado'].sum()} procedimento(s) marcados como deletados "
        f"foram excluídos do cálculo de faturamento.\n"
    )

    def montar_resumo(df, coluna_id, coluna_nome, rotulo):
        resumo = (
            df.groupby([coluna_id, coluna_nome])
            .agg(
                Faturamento_Total=("Valor_Final", "sum"),
                Qtd_Procedimentos=("Procedimento", "count"),
                Qtd_Tratamentos=("TreatmentId", "nunique"),
                Ticket_Medio_Procedimento=("Valor_Final", "mean"),
            )
            .reset_index()
            .rename(columns={coluna_id: "ProfessionalId", coluna_nome: "Profissional"})
            .sort_values("Faturamento_Total", ascending=False)
        )
        resumo["Ticket_Medio_Procedimento"] = resumo["Ticket_Medio_Procedimento"].round(2)
        resumo["Faturamento_Total"] = resumo["Faturamento_Total"].round(2)
        resumo.insert(0, "Visao", rotulo)
        return resumo

    # Visão 1: por profissional "dono" do orçamento
    resumo_orcamento = montar_resumo(
        df_ativos, "ProfessionalId_Orcamento", "Profissional_Orcamento", "Por Orçamento"
    )

    # Visão 2: por profissional que de fato executou cada procedimento
    resumo_executante = montar_resumo(
        df_ativos, "Dentist_PersonId_Executante", "Profissional_Executante", "Por Execução"
    )

    # Resumo por status, para você decidir depois como filtrar (ex: ao cruzar com financeiro)
    resumo_status = (
        df_ativos.groupby("Status_Orcamento")
        .agg(
            Qtd_Procedimentos=("Procedimento", "count"),
            Valor_Total=("Valor_Final", "sum"),
        )
        .reset_index()
        .sort_values("Valor_Total", ascending=False)
    )
    resumo_status["Valor_Total"] = resumo_status["Valor_Total"].round(2)

    return df_detalhado, resumo_orcamento, resumo_executante, resumo_status


def gerar_excel(df_detalhado, resumo_orcamento, resumo_executante, resumo_status, data_inicio, data_fim):
    """
    Gera o Excel com dashboard UCJ + 3 abas de dados.

    Identidade visual UCJ:
      Bordô principal  #871F1B
      Bordô escuro     #5C1F1D
      Rosa claro       #FAEDEC
      Rosa médio       #F5D4D3
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Paleta UCJ ───────────────────────────────────────────────────────────
    BORDE     = "871F1B"
    BORDE_ESC = "5C1F1D"
    BORDE_CLR = "FAEDEC"
    BORDE_MCL = "F5D4D3"
    BRANCO    = "FFFFFF"
    CINZA_ESC = "2C2C2C"
    CINZA_MED = "6C757D"
    VERDE_OK  = "1A7A1A"
    AMARELO   = "7A5C00"

    def _fill(h):
        return PatternFill("solid", start_color=h, end_color=h)

    def _font(bold=False, size=11, color=CINZA_ESC, italic=False):
        return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)

    def _center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    def _border_thin():
        s = Side(style="thin", color="E0E0E0")
        return Border(left=s, right=s, top=s, bottom=s)

    def _fmt_brl(v):
        return f"R$ {v:,.0f}".replace(",", ".")

    def _fmt_brl2(v):
        s = f"R$ {v:,.2f}"
        parts = s.split(".")
        parts[0] = parts[0].replace(",", ".")
        return ",".join(parts)

    # ── Pré-processar dados ──────────────────────────────────────────────────
    df_aprov = df_detalhado[df_detalhado["Status_Orcamento"] == "APPROVED"]
    aprov_por = (
        df_aprov.groupby("Profissional_Executante")
        .agg(Fat_Aprovado=("Valor_Final", "sum"),
             Proc_Aprovado=("Procedimento", "count"))
        .reset_index()
        .rename(columns={"Profissional_Executante": "Profissional"})
    )

    df_rank = resumo_orcamento.drop(columns=["Visao"]).copy()
    df_rank = df_rank.merge(aprov_por, on="Profissional", how="left")
    df_rank["Fat_Aprovado"]  = df_rank["Fat_Aprovado"].fillna(0).astype(int)
    df_rank["Proc_Aprovado"] = df_rank["Proc_Aprovado"].fillna(0).astype(int)
    df_rank["Taxa_Aprov_Pct"] = (
        df_rank["Fat_Aprovado"] / df_rank["Faturamento_Total"].replace(0, 1) * 100
    ).round(1)
    df_rank = df_rank.sort_values("Faturamento_Total", ascending=False).reset_index(drop=True)

    total_fat   = int(df_rank["Faturamento_Total"].sum())
    total_aprov = int(df_rank["Fat_Aprovado"].sum())
    total_proc  = int(df_rank["Qtd_Procedimentos"].sum())
    n_dent      = len(df_rank)
    ticket      = round(df_rank["Faturamento_Total"].sum() / max(df_rank["Qtd_Procedimentos"].sum(), 1), 2)
    total_val_s = int(resumo_status["Valor_Total"].sum())

    n = len(df_rank)
    t1 = max(1, n // 3)
    t2 = max(2, 2 * n // 3)

    # ── Workbook ─────────────────────────────────────────────────────────────
    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    #  ABA 1 — DASHBOARD
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90

    # larguras
    col_w = [1, 6, 30, 5, 14, 3, 14, 3, 14, 3, 14, 3, 14, 1]
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # alturas base
    base_rows = {1:6, 2:50, 3:6, 4:20, 5:24, 6:6, 7:6, 8:22, 9:36, 19:6, 20:22, 31:6}
    for r, h in base_rows.items():
        ws.row_dimensions[r].height = h
    for r in range(10, 10 + n):
        ws.row_dimensions[r].height = 34
    for r in range(21, 21 + len(resumo_status) + 1):
        ws.row_dimensions[r].height = 30

    # ── CABEÇALHO ─────────────────────────────────────────────────────────
    for col in range(2, 14):
        ws.cell(row=2, column=col).fill = _fill(BORDE)

    ws.merge_cells("B2:C2")
    c = ws["B2"]
    c.value = "╱ UCJ"
    c.font  = _font(bold=True, size=18, color=BORDE_MCL)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.merge_cells("D2:J2")
    c = ws["D2"]
    c.value = "ANÁLISE DE FATURAMENTO POR DENTISTA"
    c.font  = _font(bold=True, size=20, color=BRANCO)
    c.alignment = _center()

    ws.merge_cells("K2:M2")
    c = ws["K2"]
    c.value = f"UFMG Consultoria Júnior  |  {data_inicio} a {data_fim}"
    c.font  = _font(italic=True, size=9, color=BORDE_CLR)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # ── KPIs ──────────────────────────────────────────────────────────────
    kpis = [
        ("FATURAMENTO TOTAL",  _fmt_brl(total_fat)),
        ("FAT. APROVADO",      _fmt_brl(total_aprov)),
        ("PROCEDIMENTOS",      str(total_proc)),
        ("DENTISTAS ATIVOS",   str(n_dent)),
        ("TICKET MÉDIO",       _fmt_brl2(ticket)),
    ]
    kpi_spans = [(2,4), (5,6), (7,8), (9,10), (11,13)]

    for (label, valor), (c1, c2) in zip(kpis, kpi_spans):
        cl1 = get_column_letter(c1)
        cl2 = get_column_letter(c2)

        ws.merge_cells(f"{cl1}4:{cl2}4")
        cell = ws[f"{cl1}4"]
        cell.value = label
        cell.font  = _font(bold=True, size=7, color=BORDE_ESC)
        cell.fill  = _fill("F0E4E3")
        cell.alignment = _center()
        for cc in range(c1+1, c2+1):
            ws.cell(row=4, column=cc).fill = _fill("F0E4E3")

        ws.merge_cells(f"{cl1}5:{cl2}5")
        cell = ws[f"{cl1}5"]
        cell.value = valor
        cell.font  = _font(bold=True, size=15, color=BORDE)
        cell.fill  = _fill(BRANCO)
        cell.alignment = _center()
        cell.border = Border(bottom=Side(style="medium", color=BORDE))
        for cc in range(c1+1, c2+1):
            ws.cell(row=5, column=cc).fill = _fill(BRANCO)

    for col in range(2, 14):
        ws.cell(row=6, column=col).fill = _fill(BORDE_MCL)

    # ── TÍTULO RANKING ─────────────────────────────────────────────────────
    ws.merge_cells("B8:M8")
    c = ws["B8"]
    c.value = "  RANKING DE DENTISTAS — DECISÃO DE MANUTENÇÃO"
    c.font  = _font(bold=True, size=11, color=BRANCO)
    c.fill  = _fill(BORDE_ESC)
    c.alignment = _left()
    for col in range(3, 14):
        ws.cell(row=8, column=col).fill = _fill(BORDE_ESC)

    # ── CABEÇALHO TABELA ───────────────────────────────────────────────────
    hdrs = [
        ("#",               (2, 2)),
        ("DENTISTA  ·  RECOMENDAÇÃO", (3, 5)),
        ("FAT. TOTAL (R$)", (6, 7)),
        ("FAT. APROVADO",   (8, 9)),
        ("TAXA APROV.",     (10, 10)),
        ("PROCEDIMENTOS",   (11, 11)),
        ("TRATAMENTOS",     (12, 12)),
        ("TICKET MÉDIO",    (13, 13)),
    ]
    for label, (c1, c2) in hdrs:
        ws.merge_cells(f"{get_column_letter(c1)}9:{get_column_letter(c2)}9")
        cell = ws[f"{get_column_letter(c1)}9"]
        cell.value = label
        cell.font  = _font(bold=True, size=8, color=BRANCO)
        cell.fill  = _fill(BORDE)
        cell.alignment = _center()
        cell.border = Border(right=Side(style="thin", color="C0C0C0"))
        for cc in range(c1+1, c2+1):
            ws.cell(row=9, column=cc).fill = _fill(BORDE)

    # ── LINHAS DOS DENTISTAS ───────────────────────────────────────────────
    for i, row in df_rank.iterrows():
        r = 10 + i
        if i < t1:
            bg, rec, rec_cor = "E8F5E9", "▲ MANTER",  VERDE_OK
        elif i < t2:
            bg, rec, rec_cor = "FFF9E6", "◉ AVALIAR", AMARELO
        else:
            bg, rec, rec_cor = "FDECEA", "▼ REVISAR", BORDE

        dentista_val = f"{row['Profissional'].title()}"

        dados_linha = [
            (f"{i+1}",                         (2, 2),   _center(), _font(bold=True, size=10, color=BORDE)),
            (dentista_val,                      (3, 4),   _left(),   _font(bold=True, size=9,  color=CINZA_ESC)),
            (rec,                               (5, 5),   _center(), _font(bold=True, size=8,  color=rec_cor)),
            (_fmt_brl(row["Faturamento_Total"]), (6, 7),  _center(), _font(bold=True, size=9,  color=BORDE_ESC)),
            (_fmt_brl(row["Fat_Aprovado"]),     (8, 9),   _center(), _font(size=9, color=VERDE_OK if row["Fat_Aprovado"]>0 else CINZA_MED)),
            (f"{row['Taxa_Aprov_Pct']:.1f}%",  (10, 10), _center(), _font(size=9)),
            (str(row["Qtd_Procedimentos"]),     (11, 11), _center(), _font(size=9)),
            (str(row["Qtd_Tratamentos"]),       (12, 12), _center(), _font(size=9)),
            (_fmt_brl2(row["Ticket_Medio_Procedimento"]), (13, 13), _center(), _font(size=9)),
        ]

        for val, (c1, c2), aln, fnt in dados_linha:
            ws.merge_cells(f"{get_column_letter(c1)}{r}:{get_column_letter(c2)}{r}")
            cell = ws[f"{get_column_letter(c1)}{r}"]
            cell.value = val
            cell.font  = fnt
            cell.fill  = _fill(bg)
            cell.alignment = aln
            cell.border = Border(
                bottom=Side(style="thin", color="E0E0E0"),
                right=Side(style="thin",  color="E0E0E0"),
            )
            for cc in range(c1+1, c2+1):
                ws.cell(row=r, column=cc).fill = _fill(bg)

    # ── LEGENDA ────────────────────────────────────────────────────────────
    leg_row = 10 + n
    ws.row_dimensions[leg_row].height = 16
    ws.merge_cells(f"B{leg_row}:M{leg_row}")
    c = ws[f"B{leg_row}"]
    c.value = ("  ▲ MANTER = top 1/3 por faturamento    "
               "◉ AVALIAR = faixa média    "
               "▼ REVISAR = bottom 1/3 — analisar custo × benefício")
    c.font  = _font(italic=True, size=8, color=BORDE_ESC)
    c.fill  = _fill(BORDE_CLR)
    c.alignment = _left()
    for col in range(3, 14):
        ws.cell(row=leg_row, column=col).fill = _fill(BORDE_CLR)

    # ── TÍTULO STATUS ──────────────────────────────────────────────────────
    st_row = leg_row + 2
    ws.row_dimensions[st_row].height = 22
    ws.merge_cells(f"B{st_row}:M{st_row}")
    c = ws[f"B{st_row}"]
    c.value = "  STATUS DOS ORÇAMENTOS NO PERÍODO"
    c.font  = _font(bold=True, size=11, color=BRANCO)
    c.fill  = _fill(BORDE_ESC)
    c.alignment = _left()
    for col in range(3, 14):
        ws.cell(row=st_row, column=col).fill = _fill(BORDE_ESC)

    # cabeçalho status
    sh_row = st_row + 1
    ws.row_dimensions[sh_row].height = 20
    for label, (c1, c2) in [
        ("STATUS",              (2, 4)),
        ("QTD. PROCEDIMENTOS",  (5, 7)),
        ("VALOR TOTAL (R$)",    (8, 10)),
        ("% DO TOTAL",          (11, 13)),
    ]:
        ws.merge_cells(f"{get_column_letter(c1)}{sh_row}:{get_column_letter(c2)}{sh_row}")
        cell = ws[f"{get_column_letter(c1)}{sh_row}"]
        cell.value = label
        cell.font  = _font(bold=True, size=8, color=BRANCO)
        cell.fill  = _fill(BORDE)
        cell.alignment = _center()
        for cc in range(c1+1, c2+1):
            ws.cell(row=sh_row, column=cc).fill = _fill(BORDE)

    # dados de status
    stat_colors = {
        "APPROVED": ("E8F5E9", VERDE_OK, "✔ APROVADO"),
        "OPEN":     ("FFF9E6", AMARELO,  "⏳ EM ABERTO"),
        "REJECTED": ("FDECEA", BORDE,    "✖ REJEITADO"),
    }
    for i, srow in resumo_status.reset_index(drop=True).iterrows():
        r = sh_row + 1 + i
        ws.row_dimensions[r].height = 28
        status = srow["Status_Orcamento"]
        bg, fc, label = stat_colors.get(status, ("FFFFFF", CINZA_ESC, status))
        pct = round(srow["Valor_Total"] / max(total_val_s, 1) * 100, 1)

        for val, (c1, c2) in [
            (label,                         (2, 4)),
            (str(srow["Qtd_Procedimentos"]), (5, 7)),
            (_fmt_brl(srow["Valor_Total"]), (8, 10)),
            (f"{pct:.1f}%",                 (11, 13)),
        ]:
            ws.merge_cells(f"{get_column_letter(c1)}{r}:{get_column_letter(c2)}{r}")
            cell = ws[f"{get_column_letter(c1)}{r}"]
            cell.value = val
            cell.font  = _font(bold=True, size=10, color=fc)
            cell.fill  = _fill(bg)
            cell.alignment = _center()
            cell.border = Border(bottom=Side(style="thin", color="E0E0E0"))
            for cc in range(c1+1, c2+1):
                ws.cell(row=r, column=cc).fill = _fill(bg)

    # ── RODAPÉ ─────────────────────────────────────────────────────────────
    rod_row = sh_row + 1 + len(resumo_status) + 1
    ws.row_dimensions[rod_row].height = 16
    ws.merge_cells(f"B{rod_row}:M{rod_row}")
    c = ws[f"B{rod_row}"]
    c.value = "  UCJ — UFMG Consultoria Júnior  |  Dados via API Clinicorp"
    c.font  = _font(italic=True, size=8, color=BRANCO)
    c.fill  = _fill(BORDE)
    c.alignment = _left()
    for col in range(3, 14):
        ws.cell(row=rod_row, column=col).fill = _fill(BORDE)

    # ══════════════════════════════════════════════════════════════════════════
    #  HELPER: formatar abas de dados
    # ══════════════════════════════════════════════════════════════════════════
    def _formatar_aba(ws_d, df, titulo):
        ws_d.sheet_view.showGridLines = False
        ncols = len(df.columns)

        ws_d.row_dimensions[1].height = 28
        ws_d.merge_cells(f"A1:{get_column_letter(ncols)}1")
        h = ws_d["A1"]
        h.value = titulo
        h.font  = _font(bold=True, size=13, color=BRANCO)
        h.fill  = _fill(BORDE)
        h.alignment = _left()

        ws_d.row_dimensions[2].height = 13
        ws_d.merge_cells(f"A2:{get_column_letter(ncols)}2")
        t = ws_d["A2"]
        t.value = "UCJ — UFMG Consultoria Júnior"
        t.font  = _font(italic=True, size=8, color=BORDE_ESC)
        t.fill  = _fill(BORDE_CLR)
        t.alignment = _left()

        ws_d.row_dimensions[3].height = 22
        for ci, col_name in enumerate(df.columns, 1):
            cell = ws_d.cell(row=3, column=ci)
            cell.value = col_name.replace("_", " ").upper()
            cell.font  = _font(bold=True, size=9, color=BRANCO)
            cell.fill  = _fill(BORDE)
            cell.alignment = _center()
            cell.border = _border_thin()

        for ri, (_, row) in enumerate(df.iterrows(), 4):
            bg = BRANCO if (ri - 4) % 2 == 0 else "F9F3F2"
            ws_d.row_dimensions[ri].height = 17
            for ci, val in enumerate(row, 1):
                cell = ws_d.cell(row=ri, column=ci)
                cell.value = val
                cell.font  = _font(size=9, color=CINZA_ESC)
                cell.fill  = _fill(bg)
                cell.alignment = _center()
                cell.border = _border_thin()

        for ci, col_name in enumerate(df.columns, 1):
            max_len = max(
                len(str(col_name)),
                df[col_name].astype(str).str.len().max() if len(df) else 0,
            )
            ws_d.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 38)

        ws_d.freeze_panes = "A4"

    # ══════════════════════════════════════════════════════════════════════════
    #  ABA 2 — RESUMO POR DENTISTA
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Resumo por Dentista")
    df_show = df_rank[["Profissional", "Faturamento_Total", "Fat_Aprovado",
                        "Taxa_Aprov_Pct", "Qtd_Procedimentos",
                        "Qtd_Tratamentos", "Ticket_Medio_Procedimento"]].copy()
    df_show.columns = ["Profissional", "Fat_Total_R$", "Fat_Aprovado_R$",
                       "Taxa_Aprovacao_%", "Qtd_Procedimentos",
                       "Qtd_Tratamentos", "Ticket_Medio_R$"]
    _formatar_aba(ws2, df_show, "RESUMO POR DENTISTA — VISÃO CONSOLIDADA")

    # ══════════════════════════════════════════════════════════════════════════
    #  ABA 3 — ANÁLISE DE PROCEDIMENTOS
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Analise de Procedimentos")
    df_atv = df_detalhado[~df_detalhado["Deletado"]].copy()

    # Pivot: procedimento x status (qtd e valor)
    pv = (
        df_atv.groupby(["Procedimento", "Status_Orcamento"])
        .agg(Qtd=("Valor_Final", "count"), Valor=("Valor_Final", "sum"))
        .reset_index()
    )
    pivot = pv.pivot_table(
        index="Procedimento",
        columns="Status_Orcamento",
        values=["Qtd", "Valor"],
        fill_value=0,
    )
    pivot.columns = [f"{stat}_{metric}" for metric, stat in pivot.columns]
    pivot = pivot.reset_index()

    # Garantir colunas mesmo quando algum status não existe
    for col in ["APPROVED_Qtd", "APPROVED_Valor", "OPEN_Qtd", "OPEN_Valor",
                "REJECTED_Qtd", "REJECTED_Valor"]:
        if col not in pivot.columns:
            pivot[col] = 0

    # Dentistas que executam cada procedimento (lista e contagem)
    dent_por_proc = (
        df_atv.groupby("Procedimento")["Profissional_Executante"]
        .agg(N_Dentistas="nunique", Dentistas=lambda x: ", ".join(sorted(x.unique())))
        .reset_index()
    )

    # Ticket médio aprovado por procedimento
    ticket_proc = (
        df_atv[df_atv["Status_Orcamento"] == "APPROVED"]
        .groupby("Procedimento")["Valor_Final"]
        .mean()
        .round(2)
        .reset_index()
        .rename(columns={"Valor_Final": "Ticket_Medio_Aprovado"})
    )

    df_proc_final = (
        pivot
        .merge(dent_por_proc, on="Procedimento", how="left")
        .merge(ticket_proc,   on="Procedimento", how="left")
    )
    df_proc_final["Ticket_Medio_Aprovado"] = df_proc_final["Ticket_Medio_Aprovado"].fillna(0)
    df_proc_final["Total_Orcado"] = (
        df_proc_final["APPROVED_Qtd"] +
        df_proc_final["OPEN_Qtd"] +
        df_proc_final["REJECTED_Qtd"]
    ).astype(int)
    df_proc_final["Valor_Total_Aprovado"] = df_proc_final["APPROVED_Valor"].astype(int)
    df_proc_final["CRITICO"] = df_proc_final["N_Dentistas"].apply(
        lambda n: "SIM" if n == 1 else ""
    )
    df_proc_final = df_proc_final.sort_values("Valor_Total_Aprovado", ascending=False)

    df_proc_excel = df_proc_final[[
        "Procedimento", "Total_Orcado",
        "APPROVED_Qtd", "OPEN_Qtd", "REJECTED_Qtd",
        "Valor_Total_Aprovado", "Ticket_Medio_Aprovado",
        "N_Dentistas", "CRITICO", "Dentistas",
    ]].copy()
    df_proc_excel.columns = [
        "Procedimento", "Total Orcado",
        "Qtd Aprovado", "Qtd Em Aberto", "Qtd Rejeitado",
        "Valor Aprovado R$", "Ticket Medio Aprov R$",
        "N Dentistas", "Critico (1 dentista)", "Dentistas Executantes",
    ]
    _formatar_aba(ws3, df_proc_excel, "ANALISE DE PROCEDIMENTOS — VOLUME, VALOR E CRITICIDADE")

    # Destacar linhas críticas em rosa
    for ri in range(4, 4 + len(df_proc_excel)):
        critico_val = ws3.cell(row=ri, column=9).value
        if critico_val == "SIM":
            for ci in range(1, len(df_proc_excel.columns) + 1):
                ws3.cell(row=ri, column=ci).fill = _fill(BORDE_MCL)
                ws3.cell(row=ri, column=ci).font = _font(bold=True, size=9, color=BORDE_ESC)

    # ══════════════════════════════════════════════════════════════════════════
    #  ABA 4 — PROCEDIMENTO x DENTISTA (CRUZAMENTO)
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Proc x Dentista")
    df_cross = (
        df_atv.groupby(["Procedimento", "Profissional_Executante", "Status_Orcamento"])
        .agg(Qtd=("Valor_Final", "count"), Valor=("Valor_Final", "sum"))
        .reset_index()
    )
    pv2 = df_cross.pivot_table(
        index=["Procedimento", "Profissional_Executante"],
        columns="Status_Orcamento",
        values=["Qtd", "Valor"],
        fill_value=0,
    )
    pv2.columns = [f"{stat}_{metric}" for metric, stat in pv2.columns]
    pv2 = pv2.reset_index()
    for col in ["APPROVED_Qtd", "APPROVED_Valor", "OPEN_Qtd", "OPEN_Valor",
                "REJECTED_Qtd", "REJECTED_Valor"]:
        if col not in pv2.columns:
            pv2[col] = 0
    pv2["Total_Qtd"] = (pv2["APPROVED_Qtd"] + pv2["OPEN_Qtd"] + pv2["REJECTED_Qtd"]).astype(int)
    pv2 = pv2.sort_values(["Procedimento", "APPROVED_Valor"], ascending=[True, False])
    pv2_excel = pv2[[
        "Procedimento", "Profissional_Executante",
        "Total_Qtd", "APPROVED_Qtd", "OPEN_Qtd", "REJECTED_Qtd",
        "APPROVED_Valor", "OPEN_Valor", "REJECTED_Valor",
    ]].copy()
    pv2_excel.columns = [
        "Procedimento", "Dentista",
        "Total Qtd", "Qtd Aprovado", "Qtd Em Aberto", "Qtd Rejeitado",
        "Valor Aprovado R$", "Valor Em Aberto R$", "Valor Rejeitado R$",
    ]
    _formatar_aba(ws4, pv2_excel, "PROCEDIMENTO X DENTISTA — CRUZAMENTO DETALHADO")

    # ══════════════════════════════════════════════════════════════════════════
    #  ABA 5 — RESUMO POR STATUS
    # ══════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Resumo por Status")
    _formatar_aba(ws5, resumo_status, "RESUMO POR STATUS DE ORÇAMENTO")

    # ══════════════════════════════════════════════════════════════════════════
    #  ABA 6 — DETALHADO
    # ══════════════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Detalhado")
    _formatar_aba(ws6, df_detalhado, "DADOS DETALHADOS — PROCEDIMENTOS POR ORÇAMENTO")

    # ── Salvar ────────────────────────────────────────────────────────────────
    wb.save(OUTPUT_FILE)

    print(f"\nPlanilha gerada com sucesso: {OUTPUT_FILE}")
    print(f"Período analisado: {data_inicio} a {data_fim}")
    print(
        "\nAbas geradas:\n"
        "  - Dashboard: ranking de dentistas com semáforo UCJ + KPIs + status\n"
        "  - Resumo por Dentista: tabela consolidada por profissional\n"
        "  - Analise de Procedimentos: volume, valor, criticidade por procedimento\n"
        "  - Proc x Dentista: cruzamento procedimento x dentista x status\n"
        "  - Resumo por Status: APPROVED / OPEN / REJECTED\n"
        "  - Detalhado: uma linha por procedimento"
    )


# =========================================================
# EXECUÇÃO
# =========================================================

if __name__ == "__main__":
    data_inicio, data_fim = pedir_periodo()
    registros = buscar_estimates(data_inicio, data_fim)
    df_detalhado, resumo_orcamento, resumo_executante, resumo_status = processar_dados(registros)
    gerar_excel(df_detalhado, resumo_orcamento, resumo_executante, resumo_status, data_inicio, data_fim)
